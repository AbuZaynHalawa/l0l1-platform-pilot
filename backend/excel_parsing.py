"""[PO Lifecycle] Parser for the long-lead-items Excel owners already
maintain (e.g. "EST-1523 Long lead Items_WCt.xlsx"). Extracts a flat preview
list of items from the "Local Material" and "Imported Material" sheets --
the "Subcontractor" sheet is always ignored, since it only ever holds the
fixed "Design and Engineering" row (the Consultancy PO default, tracked
separately, not a real per-item list).

Real template shape (confirmed by direct inspection of a live estimate
file): a header/estimate-info block at the top of each sheet, then a row
whose column B reads "ITEM DESCRIPTION" (case/position vary between
sheets -- row 10 in one, row 9 in another -- so it's located by scanning,
not assumed fixed), a blank sub-header row right after it, then data rows
until column B goes empty again (a totals row follows with only the cost
columns populated, no item name).
"""
from io import BytesIO

from openpyxl import load_workbook

_SHEET_NAMES = {"local material", "imported material"}


def parse_long_lead_workbook(content: bytes) -> list[dict]:
    """Returns preview rows only -- never call this and commit its output
    directly. The caller must show these to the owner for confirm/edit
    first (a spreadsheet that doesn't match the expected layout should
    surface as "found 0 items", not silently create a wrong registry).
    """
    wb = load_workbook(BytesIO(content), data_only=True)
    out: list[dict] = []
    for ws in wb.worksheets:
        if ws.title.strip().lower() not in _SHEET_NAMES:
            continue  # skips "Subcontractor" and anything unexpected
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=30):
            cell_b = row[1].value if len(row) > 1 else None
            if str(cell_b or "").strip().upper() == "ITEM DESCRIPTION":
                header_row = row[0].row
                break
        if header_row is None:
            continue
        r = header_row + 2  # header row, one blank-in-B sub-header row, then data
        while True:
            name = ws.cell(row=r, column=2).value
            if name is None or str(name).strip() == "":
                break
            out.append({
                "sheet": ws.title.strip(),
                "name": str(name).strip(),
                "qty": ws.cell(row=r, column=3).value,
                "unit": ws.cell(row=r, column=4).value,
                "supplier": ws.cell(row=r, column=7).value,
                "delivery_est": ws.cell(row=r, column=16).value,
            })
            r += 1
    return out
