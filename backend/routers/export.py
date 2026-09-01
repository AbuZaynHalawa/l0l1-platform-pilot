"""[Request 3]: one generic, reusable "Export to Excel" endpoint every
table in the platform shares, instead of a bespoke backend export per
table. The frontend already has whatever rows a table is currently
showing (filtered/sorted client-side via installExcelHeader) sitting in
memory -- it just POSTs that same {columns, rows} shape here and gets
back a real, styled .xlsx (bold colored header, zebra striping, borders,
autosized columns), not a raw data dump. See app.js's exportTableToExcel()
for the frontend half.
"""
import io
import re
from datetime import datetime

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/export", tags=["export"])

# Algihaz brand red (matches --algihaz-red in styles.css) for the title
# band; a near-black for the header row -- the same "dark header, light
# zebra body" look every table in the app itself already uses.
_BRAND_RED = "D71920"
_HEADER_DARK = "252A34"
_ZEBRA = "F4F5F8"
_BORDER_COLOR = "DCDFE6"


class ExportColumn(BaseModel):
    key: str
    label: str


class ExportRequest(BaseModel):
    title: str
    columns: list[ExportColumn]
    rows: list[dict]
    sheet_name: str = "Data"


def _safe_filename(title: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9 _-]", "", title).strip().replace(" ", "_")
    return (cleaned or "export") + ".xlsx"


@router.post("/xlsx")
def export_xlsx(payload: ExportRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\[\]:\*\?/\\]", "", payload.sheet_name or "Data")[:31] or "Data"

    ncols = max(len(payload.columns), 1)
    thin = Side(style="thin", color=_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title band + export timestamp, both merged across every column so
    # they read as a real report header, not a stray first cell.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value=payload.title)
    title_cell.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=_BRAND_RED)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub_cell = ws.cell(row=2, column=1, value="Algihaz — Project Readiness (L0/L1) Platform  ·  Exported "
                        + datetime.now().strftime("%d %b %Y, %H:%M"))
    sub_cell.font = Font(name="Calibri", size=9.5, italic=True, color="6B7280")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    header_row = 4
    for ci, col in enumerate(payload.columns, start=1):
        c = ws.cell(row=header_row, column=ci, value=col.label)
        c.font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_HEADER_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    zebra_fill = PatternFill("solid", fgColor=_ZEBRA)
    col_widths = [len(col.label) for col in payload.columns]
    for ri, row in enumerate(payload.rows, start=header_row + 1):
        is_zebra = (ri - header_row) % 2 == 0
        for ci, col in enumerate(payload.columns, start=1):
            val = row.get(col.key, "")
            if val is None:
                val = ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=10.5, color="1F2430")
            c.border = border
            c.alignment = Alignment(vertical="center", indent=1, wrap_text=False)
            if is_zebra:
                c.fill = zebra_fill
            col_widths[ci - 1] = max(col_widths[ci - 1], len(str(val)))

    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 3, 11), 52)

    if not payload.rows:
        ws.cell(row=header_row + 1, column=1, value="No rows to export.").font = Font(name="Calibri", italic=True, color="9AA0AC")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_safe_filename(payload.title)}"'},
    )
